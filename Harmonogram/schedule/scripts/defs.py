import json

from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = load_workbook(
    filename="D:\Python\Projekt_harmonogramu_JMP\Harmonogram prac JMP (1).xlsx",
    data_only=True)

sheet = wb["2022"]
list_of_values = [i for i in sheet[1]]

def cell_value(searching_name, counting):
    for i in list_of_values:
        if i.value == searching_name:
            letter = get_column_letter(i.column)
            return sheet[f"{letter}{counting}"].value
    return None


def data_validator(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    else:
        return None


def get_month(month: str):
    dict_of_months = {
        "Styczeń": 1,
        "Luty": 2,
        "Marzec": 3,
        "Kwiecień": 4,
        "Maj": 5,
        "Czerwiec": 6,
        "Lipiec": 7,
        "Sierpień": 8,
        "Wrzesień": 9,
        "Październik": 10,
        "Listopad": 11,
        "Grudzień": 12,
    }
    if month != None and month in dict_of_months.keys():
        month = month.strip()
        return dict_of_months[month]
    else:
        return None


def get_street_or_number(string=str, position=int):
    if string:
        values = string.split()
        if len(values) == 1:
            return str(*values) if position == 0 else None
        elif len(values) == 2 and values[-1].isdigit():
            return values[position]
        elif len(values) == 2 and not values[-1].isdigit():
            return "".join(values) if position == 0 else None
        else:
            if position == 1:
                return values[-1]
            else:
                return " ".join(values[:-1])

    else:
        return None


def data_changer(iterator):
    day = cell_value("Start montażu - dostawa listy zmiennej i mebli", iterator)
    year = int(sheet.title)
    month = get_month(cell_value("Miesiac montażu", iterator))
    print(type(day), type(month), iterator)
    if not all(isinstance(i, int) for i in [day, year, month]):
        return None
    else:
        return datetime(year, month, day).strftime("%Y-%m-%d %H:%M:%S")



def create_json():
    list_of_data = []
    for i in range(2, sheet.max_row):
        x = {
            "model": "schedule.store",
            "pk": i,
            "fields": {
                "store_number": f'{cell_value("Nr sklepu", i)}',
                "store_city": f'{cell_value("Miasto dostawy", i)}',
                "store_street": f'{get_street_or_number(cell_value("Ulica dostawy", i), 0)}',
                "store_street_number": f'{get_street_or_number(cell_value("Ulica dostawy", i), 1)}',
                "zip_code": f'{cell_value("Kod pocztowy", i)}',
                "date_start_installation": data_changer(i),
                "date_opening": data_validator(cell_value("Data otwarcia sklepu", i)),
                "date_disassembling": data_validator(cell_value("Demontaż", i)),
                "region": f'{cell_value("Region", i)}'
            }
        }
        list_of_data.append(x)
        with open("json_file.json", "w", encoding='utf-8') as outfile:
            json.dump(list_of_data, outfile, ensure_ascii=False)
        file = open("json_file.json", mode="r")
    return file


print(cell_value("Miasto dostawy", 5))
print(create_json())


# def in_one_month(date_opening):
#     now = datetime.now()
#     return

# date = datetime.now().date() + timedelta(days=30)

data_1 = 2023-11-21
data_2 = 2023-10-22
data_3 = 2023-11-23

def if_data(data_1, data_2, data_3):
    return data_1 >= data_2 >= data_3

print(if_data(data_1, data_2, data_3))